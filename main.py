import fitz

import openpyxl

pdf = fitz.open('D:\\GitProject\\addressbok_pdf_excel\\provider_results.pdf')

x = ''

for y in range(2, 622):
    data = pdf.load_page(y).get_text()
    x = x + data

for inpt in range(150):

    wb = openpyxl.load_workbook('D:\\GitProject\\addressbok_pdf_excel\\sheetpdftoexcel.xlsx')
    inputsheet = wb['Sheet1']
    inputrow = inpt + 1

    start = x.index('Name:')
    if 'mile' in x[:30]:
        cut = x.index('mile') + 5
        title = x[cut:start]
    elif 'miles' in x[0:30]:
        cut = x.index('miles') + 6
        title = x[cut:start]
    else:
        title = x[:start]

    inputsheet.cell(row=inputrow + 1, column=1, value=title)
    wb.save("sheetpdftoexcel.xlsx")

    a = x.index('Specialty:')
    name = x[start:a]
    name = name.replace('Name: ', '')
    inputsheet.cell(row=inputrow + 1, column=2, value=name)
    wb.save("sheetpdftoexcel.xlsx")

    b = x.index('NPI:')
    specialty = x[a:b]
    specialty = specialty.replace('Specialty: ', '')
    inputsheet.cell(row=inputrow + 1, column=3, value=specialty)
    wb.save("sheetpdftoexcel.xlsx")

    c = x.index('Location:')
    NPI = x[b:c]
    NPI = NPI.replace('NPI: ', '')
    inputsheet.cell(row=inputrow + 1, column=4, value=NPI)
    wb.save("sheetpdftoexcel.xlsx")

    d = x.index('Address:')
    location = x[c:d]
    location = location.replace('Location: ', '')
    inputsheet.cell(row=inputrow + 1, column=5, value=location)
    wb.save("sheetpdftoexcel.xlsx")

    e = x.index('Distance:')
    address = x[d:e]
    address = address.replace('Address: ', '')
    inputsheet.cell(row=inputrow + 1, column=6, value=address)
    wb.save("sheetpdftoexcel.xlsx")

    f = x.index('Phone:')
    distance = x[e:f]
    distance = distance.replace('Distance: ', '')
    inputsheet.cell(row=inputrow + 1, column=7, value=distance)
    wb.save("sheetpdftoexcel.xlsx")

    end = x.index('Patients:') + 13

    x = x[end:]

    wb.save("sheetpdftoexcel.xlsx")


